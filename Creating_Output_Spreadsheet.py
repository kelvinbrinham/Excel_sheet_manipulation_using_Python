'''
Creating the output spreadsheet
'''
import pandas as pd
import numpy as np
import openpyxl as xl
import datetime
from datetime import datetime as dt
import xlsxwriter as xlw
from string import ascii_uppercase as UPPER
from openpyxl.styles import Border,Side


GS_df = pd.read_excel('Data_Formatted/EPS_CHANGE_20221028_GS_FORMATTED.xlsx')
MS_df = pd.read_excel('Data_Formatted/EPS_CHANGES_20221028_MS_FORMATTED.xlsx')
JPMCAZ_df = pd.read_excel('Data_FORMATTED/EPS_CHANGE_20221028_JPMCAZ_FORMATTED.xlsx')
INTERMONTE_df = pd.read_excel('Data_Formatted/EPS_CHANGES_20221028_INTERMONTE_FORMATTED.xlsx')

#Combine input excel sheets into one dataframe
Combined_df = pd.concat([GS_df, MS_df, JPMCAZ_df, INTERMONTE_df])
#Drop first column
Combined_df.drop([Combined_df.columns[0]], axis=1, inplace=True)

#Replace missing valyes with 'N/A'
Combined_df = Combined_df.replace(np.nan, 'N/A')
#Sort rows alphabetically by Ticker
Combined_df = Combined_df.sort_values('Ticker')

Combined_df.reset_index(inplace=True, drop=True)
#Emptying duplicate ticker cells
Tickers_set = set()
for i in range(len(Combined_df)):
    if Combined_df.iloc[i]['Ticker'] not in Tickers_set:
        Tickers_set.add(Combined_df.iloc[i]['Ticker'])
    else:
        Combined_df.at[i,'Ticker'] = np.nan
        Combined_df.at[i,'Company Name'] = np.nan

# Formatting <><><><><><><><><><><><><><><><><><><><><><><><>

# Swapping columns to match example output
Combined_df = Combined_df[['Ticker', 'Company Name', 'Broker', 'currency', 'y1_EPS chg', 'y2_EPS chg', 'y3_EPS chg',
'y1 new EPS difference to consensus', 'y2 new EPS difference to consensus', 'y2 new EPS difference to consensus'
, 'y1_PE', 'y2_PE', 'y3_PE', 'comment', 'y1 new EPS', 'y2 new EPS', 'y3 new EPS',
'y1_sales chg', 'y2_sales chg', 'y3_sales chg', 'y1_EBIT chg', 'y2_EBIT chg', 'y3_EBIT chg']]


Output_file_name = 'Data_FORMATTED/Combined_OUTPUT.xlsx'
Combined_df.to_excel(Output_file_name, sheet_name='Sheet1', startrow = 3, index = False)


Combined_wb = xl.load_workbook(Output_file_name)
Combined_ws = Combined_wb.active

#Change relevant values to percentages
#These are the letters of columns that should be in percentage format
set_of_percentage_column_letters = ['E', 'F', 'G', 'H', 'I', 'J', 'R', 'S', 'T', 'U', 'V', 'W']
#Loop over these columns and change into percentage format (ignoring missing data)
for letter in set_of_percentage_column_letters:
    for i in range(5, 87 + 1, 1): #4-87
        if Combined_ws[letter + str(i)] != 'N/A':
            Combined_ws[letter + str(i)].number_format = '0.00%'


#Changing Headers from my code names to readable names

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


Combined_ws['E3'] = 'EPS change'
Combined_ws['E3'].border = thin_border
Combined_ws.merge_cells('E3:G3')
Combined_ws['E4'] = 2023
Combined_ws['F4'] = 2024
Combined_ws['G4'] = 2025

Combined_ws['H3'] = 'new EPS vs Cons'
Combined_ws['H3'].border = thin_border
Combined_ws.merge_cells('H3:J3')
Combined_ws['H4'] = 2023
Combined_ws['I4'] = 2024
Combined_ws['J4'] = 2025

Combined_ws['K3'] = 'P/E'
Combined_ws['K3'].border = thin_border
Combined_ws.merge_cells('K3:M3')
Combined_ws['K4'] = 2023
Combined_ws['L4'] = 2024
Combined_ws['M4'] = 2025

Combined_ws['O3'] = 'new EPS'
Combined_ws['O3'].border = thin_border
Combined_ws.merge_cells('O3:Q3')
Combined_ws['O4'] = 2023
Combined_ws['P4'] = 2024
Combined_ws['Q4'] = 2025

Combined_ws['R3'] = 'Sales change'
Combined_ws['R3'].border = thin_border
Combined_ws.merge_cells('R3:T3')
Combined_ws['R4'] = 2023
Combined_ws['S4'] = 2024
Combined_ws['T4'] = 2025

Combined_ws['U3'] = 'EBIT change'
Combined_ws['U3'].border = thin_border
Combined_ws.merge_cells('U3:W3')
Combined_ws['U4'] = 2023
Combined_ws['V4'] = 2024
Combined_ws['W4'] = 2025


#Add Title
Combined_ws['A1'] = 'Analyst Estimate Changes (Exercise)'
Combined_ws['A2'] = dt.now()



Combined_wb.save(Output_file_name)
